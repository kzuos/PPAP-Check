from __future__ import annotations

import csv
import io
import json
import re
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader

from ppapcheck.models import (
    CertificateRecord,
    CharacteristicType,
    ControlPlanEntry,
    DocumentRecord,
    DocumentType,
    DrawingCharacteristic,
    EvidenceRef,
    ExtractedValue,
    ExtractionStatus,
    InspectionResult,
    MeasurementStatus,
    PfmeaEntry,
    ProcessFlowStep,
    SubmissionContext,
    SubmissionMode,
    SubmissionPackage,
)
from ppapcheck.services.document_ocr_service import DocumentOcrService


@dataclass
class TextFragment:
    text: str
    page_number: int | None = None
    section_name: str | None = None


@dataclass
class TableSection:
    section_name: str
    rows: list[list[str]]


@dataclass
class ParsedUploadDocument:
    documents: list[DocumentRecord]
    warnings: list[str]


@dataclass
class UploadBuildResult:
    package: SubmissionPackage
    warnings: list[str]


@dataclass
class PdfSection:
    document_type: DocumentType
    classification_confidence: float
    start_page: int
    end_page: int
    section_label: str
    fragments: list[TextFragment] = field(default_factory=list)


FIELD_PATTERNS: dict[str, list[tuple[re.Pattern[str], float]]] = {
    "part_number": [
        (
            re.compile(
                r"(?:^|\n)\s*(?:part\s*number|part\s*no\.?|p/?n)\s*[:#-]?\s*"
                r"([A-Z0-9][A-Z0-9 ._/\-]{2,80}?)(?=\s+(?:hardware version|part number|name|drawing number|diagnosis status|software version|version / date|identification|$)|$)",
                re.IGNORECASE,
            ),
            0.98,
        ),
    ],
    "drawing_number": [
        (
            re.compile(
                r"(?:^|\n)\s*(?:drawing\s*number|drawing\s*no\.?|dwg(?:\s*number|\s*no\.?)?)\s*[:#-]?\s*"
                r"([A-Z0-9][A-Z0-9 ._/\-]{2,80}?)(?=\s+(?:software version|name|version / date|identification|$)|$)",
                re.IGNORECASE,
            ),
            0.97,
        ),
    ],
    "revision": [
        (
            re.compile(
                r"(?:^|\n)\s*version\s*/\s*date\s*[:#-]?\s*([A-Z0-9]{1,8})(?:\s*/\s*[0-9]{1,2}[./][0-9]{1,2}[./][0-9]{2,4})?",
                re.IGNORECASE,
            ),
            0.96,
        ),
        (re.compile(r"(?:^|\n)\s*(?:revision|rev(?:ision)?\.?)\s*[:#-]?\s*([A-Z0-9][A-Z0-9._/\-]{0,12})", re.IGNORECASE), 0.92),
    ],
    "customer_name": [
        (
            re.compile(
                r"(?:^|\n)\s*customer(?!-ready)(?:\s*\(recipient\))?\s*[:#-]?\s*([^\n\r]{2,100})",
                re.IGNORECASE,
            ),
            0.9,
        ),
        (
            re.compile(
                r"\bcustomer(?!-ready)\s+([A-Z][A-Za-z0-9 .&()/\-]{2,80}?)(?=report version|delivery quantity|delivery location|batch number|order number|production location|sample weight|unloading point|$)",
                re.IGNORECASE,
            ),
            0.86,
        ),
    ],
    "supplier_name": [
        (
            re.compile(r"(?:^|\n)\s*organization\s*[:#-]?\s*([A-Z][^\n\r]{2,120})", re.IGNORECASE),
            0.92,
        ),
        (re.compile(r"(?:^|\n)\s*(?:supplier|vendor|manufacturer)\s*[:#-]?\s*([^\n\r]{2,100})", re.IGNORECASE), 0.88),
    ],
    "process_name": [
        (re.compile(r"(?:^|\n)\s*(?:process\s*name|manufacturing\s+process)\s*[:#-]?\s*([^\n\r]{2,120})", re.IGNORECASE), 0.86),
    ],
    "material": [
        (re.compile(r"(?:^|\n)\s*(?:material\b|matl\.?\b|alloy\b)\s*[:#-]?\s*([^\n\r]{2,120})", re.IGNORECASE), 0.84),
    ],
    "submission_reason": [
        (re.compile(r"(?:^|\n)\s*(?:submission\s+reason|reason\s+for\s+submission|reason\s+for\s+report\s+creation)\s*[:#-]?\s*([^\n\r]{2,140})", re.IGNORECASE), 0.86),
    ],
    "approval_status": [
        (re.compile(r"(?:^|\n)\s*(?:approval\s+status|warrant\s+status)\s*[:#-]?\s*([^\n\r]{2,80})", re.IGNORECASE), 0.82),
    ],
    "signatory": [
        (re.compile(r"(?:^|\n)\s*(?:signed\s+by|approved\s+by|signatory|prepared\s+by|name)\s*[:#-]?\s*([^\n\r]{2,100})", re.IGNORECASE), 0.8),
    ],
    "date": [
        (re.compile(r"(?:^|\n)\s*(?:date|submission\s+date|inspection\s+date)\s*[:#-]?\s*([0-9]{4}[-/][0-9]{1,2}[-/][0-9]{1,2}|[A-Za-z]{3,9}\s+\d{1,2},\s+\d{4}|\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", re.IGNORECASE), 0.8),
    ],
    "ppap_level": [
        (re.compile(r"ppap\s+level\s*[:#-]?\s*([1-5])", re.IGNORECASE), 0.9),
    ],
}

MONTH_TOKENS = (
    "ocak",
    "subat",
    "şubat",
    "mart",
    "nisan",
    "mayis",
    "mayıs",
    "haziran",
    "temmuz",
    "agustos",
    "ağustos",
    "eylul",
    "eylül",
    "ekim",
    "kasim",
    "kasım",
    "aralik",
    "aralık",
    "january",
    "february",
    "march",
    "april",
    "may",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
)

DOCUMENT_KEYWORDS: dict[DocumentType, tuple[tuple[str, float], ...]] = {
    DocumentType.PSW: (("part submission warrant", 0.95), (" psw ", 0.76), ("warrant status", 0.7)),
    DocumentType.DESIGN_RECORD: (("design record", 0.9), ("drawing number", 0.72), ("dwg", 0.66)),
    DocumentType.BALLOONED_DRAWING: (("ballooned drawing", 0.96), ("balloon", 0.72)),
    DocumentType.PFMEA: (("pfmea", 0.96), ("process failure mode", 0.92)),
    DocumentType.DFMEA: (("dfmea", 0.96), ("design failure mode", 0.92)),
    DocumentType.PROCESS_FLOW: (("process flow", 0.96), ("flow diagram", 0.84)),
    DocumentType.CONTROL_PLAN: (("control plan", 0.97),),
    DocumentType.MSA: (("measurement systems analysis", 0.94), ("gage r&r", 0.9), ("grr", 0.74)),
    DocumentType.DIMENSIONAL_RESULTS: (("dimensional results", 0.96), ("layout inspection", 0.86), ("dimensional report", 0.88)),
    DocumentType.MATERIAL_RESULTS: (("material test", 0.92), ("material results", 0.92)),
    DocumentType.PERFORMANCE_RESULTS: (("performance test", 0.92), ("performance results", 0.92)),
    DocumentType.PROCESS_STUDY: (("initial process study", 0.96), ("cpk", 0.7), ("ppk", 0.7), ("capability study", 0.88)),
    DocumentType.CUSTOMER_REQUIREMENTS: (("customer-specific requirements", 0.95), ("customer requirements", 0.88)),
    DocumentType.FAIR: (("first article inspection report", 0.98), ("fair", 0.86), ("as9102", 0.9)),
    DocumentType.MATERIAL_CERTIFICATE: (("material certificate", 0.95), ("certificate of conformance", 0.84), ("mill cert", 0.86)),
    DocumentType.SPECIAL_PROCESS_CERTIFICATE: (("special process certificate", 0.95), ("nadcap", 0.82)),
    DocumentType.MEASUREMENT_TRACEABILITY: (("measurement traceability", 0.94), ("calibration certificate", 0.84)),
}

SPECIAL_TOKENS = ("<sc>", "special", "critical", "key characteristic", "cc")
PASS_TOKENS = {"pass", "ok", "accept", "accepted", "conform", "true"}
FAIL_TOKENS = {"fail", "reject", "rejected", "ng", "nonconforming", "false"}

HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "characteristic_id": ("characteristic id", "char id", "feature id", "item", "feature", "characteristic"),
    "balloon_number": ("balloon", "balloon number", "balloon no", "item no", "bubble"),
    "description": ("description", "characteristic description", "feature description"),
    "nominal": ("nominal", "target", "spec", "basic dimension"),
    "tolerance": ("tolerance", "tol", "+/-"),
    "unit": ("unit", "units"),
    "measured_value": ("measured", "actual", "measurement", "result value", "measured value"),
    "result": ("result", "status", "pass/fail", "accept/reject"),
    "step_id": ("step", "process step", "operation", "op", "op no", "sequence"),
    "step_name": ("step name", "operation description", "process step name", "process"),
    "control_method": ("control method", "inspection method", "method"),
    "reaction_plan": ("reaction plan", "reaction"),
    "failure_mode": ("failure mode", "mode of failure"),
    "severity": ("severity", "sev"),
    "risk_priority": ("rpn", "risk priority", "risk priority number"),
    "certificate_type": ("certificate", "cert type", "document"),
    "identifier": ("identifier", "certificate id", "cert no", "number"),
}


class UploadSubmissionBuilder:
    def __init__(self) -> None:
        self.ocr_service = DocumentOcrService()

    def build(
        self,
        files: list[tuple[str, bytes]],
        context: SubmissionContext,
    ) -> UploadBuildResult:
        warnings: list[str] = []
        documents: list[DocumentRecord] = []
        for file_name, payload in files:
            parsed = self._parse_document(file_name, payload)
            documents.extend(parsed.documents)
            warnings.extend(parsed.warnings)

        hydrated_context = self._hydrate_context(context, documents)
        package = SubmissionPackage(
            submission_id=self._submission_id(),
            submission_mode=hydrated_context.requested_submission_mode
            if hydrated_context.requested_submission_mode != SubmissionMode.UNKNOWN
            else SubmissionMode.UNKNOWN,
            context=hydrated_context,
            documents=documents,
        )
        return UploadBuildResult(package=package, warnings=warnings)

    def _parse_document(self, file_name: str, payload: bytes) -> ParsedUploadDocument:
        suffix = Path(file_name).suffix.lower()
        warnings: list[str] = []
        try:
            if suffix == ".pdf":
                fragments, tables, pdf_warnings = self._extract_pdf_content(file_name, payload)
                warnings.extend(pdf_warnings)
            else:
                fragments, tables = self._extract_content(file_name, payload, suffix)
        except Exception as exc:
            warning = f"{file_name}: parser failed ({exc}). Manual review is required."
            return ParsedUploadDocument(
                documents=[
                    DocumentRecord(
                        file_name=file_name,
                        document_type=DocumentType.UNKNOWN,
                        classification_confidence=0.25,
                        notes=[
                            "File parsing failed, so no verifiable machine extraction is available for this document.",
                        ],
                    )
                ],
                warnings=[warning],
            )
        joined_text = "\n".join(fragment.text for fragment in fragments if fragment.text).strip()

        if suffix == ".pdf":
            pdf_documents, pdf_warnings = self._parse_pdf_bundle(file_name, fragments, tables)
            warnings.extend(pdf_warnings)
            if pdf_documents:
                return ParsedUploadDocument(documents=pdf_documents, warnings=warnings)

        document_type, classification_confidence = self._classify_document(file_name, joined_text)
        metadata = self._extract_metadata(file_name, fragments)
        metadata = self._merge_filename_inference(file_name, metadata)
        structured = self._extract_structured_rows(file_name, document_type, tables)

        notes: list[str] = []
        if suffix == ".pdf" and len(joined_text) < 80:
            notes.append(
                "No reliable machine-readable PDF text was extracted. Scanned or image-based pages require OCR or manual review."
            )
            warnings.append(f"{file_name}: PDF appears image-based or text-poor; OCR or manual review is required.")
        elif suffix == ".pdf":
            sparse_pages = [
                fragment.page_number
                for fragment in fragments
                if fragment.page_number is not None and len(re.sub(r"\s+", "", fragment.text or "")) < 40
            ]
            if sparse_pages:
                page_text = ", ".join(str(page) for page in sparse_pages[:10])
                if len(sparse_pages) > 10:
                    page_text += ", ..."
                notes.append(
                    f"Text extraction is sparse on pages {page_text}; OCR or manual review is required for complete visual verification."
                )
        if document_type == DocumentType.UNKNOWN:
            notes.append(
                "Document type could not be classified confidently from file content. Validation coverage may be incomplete."
            )
        if suffix not in {".pdf", ".txt", ".csv", ".tsv", ".xlsx", ".xlsm", ".json"}:
            notes.append(
                "File type is not fully supported by the parser in this release. Only filename-based and low-confidence extraction may be available."
            )
            warnings.append(f"{file_name}: unsupported file type; parser coverage is limited.")

        document = DocumentRecord(
            file_name=file_name,
            document_type=document_type,
            classification_confidence=classification_confidence,
            metadata=metadata,
            drawing_characteristics=structured["drawing_characteristics"],
            inspection_results=structured["inspection_results"],
            process_flow_steps=structured["process_flow_steps"],
            pfmea_entries=structured["pfmea_entries"],
            control_plan_entries=structured["control_plan_entries"],
            certificates=structured["certificates"],
            notes=notes,
        )
        return ParsedUploadDocument(documents=[document], warnings=warnings)

    def _extract_content(
        self,
        file_name: str,
        payload: bytes,
        suffix: str,
    ) -> tuple[list[TextFragment], list[TableSection]]:
        if suffix == ".pdf":
            return self._extract_pdf(payload), []
        if suffix in {".txt", ".md"}:
            text = payload.decode("utf-8", errors="ignore")
            return [TextFragment(text=text, section_name="text")], []
        if suffix in {".csv", ".tsv"}:
            dialect = csv.excel_tab if suffix == ".tsv" else csv.excel
            text = payload.decode("utf-8", errors="ignore")
            rows = list(csv.reader(io.StringIO(text), dialect=dialect))
            table = TableSection(section_name="table", rows=[[self._clean_cell(cell) for cell in row] for row in rows])
            fragments = [TextFragment(text="\n".join(" | ".join(filter(None, row)) for row in table.rows), section_name="table")]
            return fragments, [table]
        if suffix in {".xlsx", ".xlsm"}:
            return self._extract_workbook(payload)
        if suffix == ".json":
            text = payload.decode("utf-8", errors="ignore")
            try:
                pretty = json.dumps(json.loads(text), ensure_ascii=True, indent=2)
            except json.JSONDecodeError:
                pretty = text
            return [TextFragment(text=pretty, section_name="json")], []
        return [TextFragment(text="", section_name="unsupported")], []

    def _extract_pdf_content(
        self,
        file_name: str,
        payload: bytes,
    ) -> tuple[list[TextFragment], list[TableSection], list[str]]:
        fragments = self._extract_pdf(payload)
        tables: list[TableSection] = []
        warnings: list[str] = []

        try:
            with pdfplumber.open(io.BytesIO(payload)) as pdf:
                plumber_fragments_by_page: dict[int, str] = {}
                for index, page in enumerate(pdf.pages, start=1):
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        plumber_fragments_by_page[index] = page_text
                    for table_index, table in enumerate(page.extract_tables(), start=1):
                        cleaned_rows = [
                            [self._clean_cell(cell) for cell in row]
                            for row in table
                            if row and any(self._clean_cell(cell) for cell in row)
                        ]
                        if not cleaned_rows:
                            continue
                        tables.append(
                            TableSection(
                                section_name=f"pdf_page_{index}_table_{table_index}",
                                rows=cleaned_rows,
                            )
                        )

                merged_fragments: list[TextFragment] = []
                for fragment in fragments:
                    page_number = fragment.page_number or 0
                    plumber_text = plumber_fragments_by_page.get(page_number, "")
                    merged_text = fragment.text
                    if len(plumber_text.strip()) > len((fragment.text or "").strip()):
                        merged_text = plumber_text
                    merged_fragments.append(
                        TextFragment(
                            text=merged_text,
                            page_number=fragment.page_number,
                            section_name=fragment.section_name,
                        )
                    )
                fragments = merged_fragments
        except Exception as exc:
            warnings.append(f"{file_name}: pdfplumber table extraction failed ({exc}).")

        low_text_pages = [
            fragment.page_number
            for fragment in fragments
            if fragment.page_number is not None and len(re.sub(r"\s+", "", fragment.text or "")) < 40
        ]
        if low_text_pages:
            page_list = ", ".join(str(page) for page in low_text_pages[:10])
            if len(low_text_pages) > 10:
                page_list += ", ..."
            if not self.ocr_service.enabled:
                warnings.append(
                    f"{file_name}: OCR fallback is not configured for text-poor PDF pages {page_list}. Manual review is required for those pages."
                )
            else:
                ocr_result = self.ocr_service.extract_pdf_pages(payload, low_text_pages[:10])
                warnings.extend(f"{file_name}: {warning}" for warning in ocr_result.warnings)
                ocr_text_by_page = {page.page_number: page.text for page in ocr_result.pages if page.text.strip()}
                if ocr_text_by_page:
                    fragments = [
                        TextFragment(
                            text=ocr_text_by_page.get(fragment.page_number, fragment.text),
                            page_number=fragment.page_number,
                            section_name=fragment.section_name,
                        )
                        for fragment in fragments
                    ]

        return fragments, tables, warnings

    def _extract_pdf(self, payload: bytes) -> list[TextFragment]:
        reader = PdfReader(io.BytesIO(payload))
        fragments: list[TextFragment] = []
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            fragments.append(TextFragment(text=text, page_number=index))
        return fragments

    def _parse_pdf_bundle(
        self,
        file_name: str,
        fragments: list[TextFragment],
        tables: list[TableSection],
    ) -> tuple[list[DocumentRecord], list[str]]:
        sections = self._build_pdf_sections(file_name, fragments, tables)
        if len(sections) <= 1:
            return [], []

        warnings = [
            f"{file_name}: bundled PDF was split into {len(sections)} virtual evidence sections for validation."
        ]
        shared_metadata = self._extract_metadata(file_name, fragments)
        documents: list[DocumentRecord] = []
        for section in sections:
            section_metadata = self._extract_metadata(file_name, section.fragments)
            section_metadata = self._merge_shared_metadata(section_metadata, shared_metadata)
            section_metadata = self._merge_filename_inference(file_name, section_metadata)
            section_tables = self._tables_for_page_span(tables, section.start_page, section.end_page)
            structured = self._extract_structured_rows(file_name, section.document_type, section_tables)
            notes = [
                f"Virtual document extracted from bundled PDF pages {self._page_span_text(section.start_page, section.end_page)}.",
                f"Detected section: {section.section_label}.",
            ]
            sparse_pages = [
                fragment.page_number
                for fragment in section.fragments
                if fragment.page_number is not None and len(re.sub(r"\s+", "", fragment.text or "")) < 40
            ]
            if sparse_pages:
                page_text = ", ".join(str(page) for page in sparse_pages)
                notes.append(
                    f"Text extraction is sparse on pages {page_text}; OCR or manual review is required for complete visual verification."
                )
            if section.document_type == DocumentType.PSW and "ppa" in section.section_label.lower():
                notes.append(
                    "VDA/PPA cover sheet was treated as PSW-equivalent submission warrant evidence."
                )

            documents.append(
                DocumentRecord(
                    file_name=self._section_file_name(file_name, section.start_page, section.end_page),
                    document_type=section.document_type,
                    classification_confidence=section.classification_confidence,
                    metadata=section_metadata,
                    drawing_characteristics=structured["drawing_characteristics"],
                    inspection_results=structured["inspection_results"],
                    process_flow_steps=structured["process_flow_steps"],
                    pfmea_entries=structured["pfmea_entries"],
                    control_plan_entries=structured["control_plan_entries"],
                    certificates=structured["certificates"],
                    notes=notes,
                )
            )

        return documents, warnings

    def _build_pdf_sections(
        self,
        file_name: str,
        fragments: list[TextFragment],
        tables: list[TableSection],
    ) -> list[PdfSection]:
        sections: list[PdfSection] = []
        mergeable_types = {
            DocumentType.DIMENSIONAL_RESULTS,
            DocumentType.MATERIAL_RESULTS,
            DocumentType.PERFORMANCE_RESULTS,
            DocumentType.OEM_SPECIFIC,
            DocumentType.ENGINEERING_CHANGE,
        }

        for fragment in fragments:
            page_number = fragment.page_number or 0
            text = fragment.text.strip()
            document_type, confidence, section_label = self._classify_pdf_page(file_name, fragment)
            table_type, table_confidence, table_label = self._classify_pdf_tables(page_number, tables)
            if table_type != DocumentType.UNKNOWN and (
                document_type == DocumentType.UNKNOWN
                or document_type in {DocumentType.DESIGN_RECORD, DocumentType.MATERIAL_CERTIFICATE}
                or table_confidence > confidence
            ):
                document_type = table_type
                confidence = table_confidence
                section_label = table_label

            if not text:
                if sections:
                    sections[-1].fragments.append(fragment)
                    sections[-1].end_page = max(sections[-1].end_page, page_number)
                continue

            if document_type == DocumentType.UNKNOWN and len(text) < 240 and sections:
                sections[-1].fragments.append(fragment)
                sections[-1].end_page = max(sections[-1].end_page, page_number)
                continue

            if (
                sections
                and sections[-1].document_type == document_type
                and document_type in mergeable_types
                and page_number == sections[-1].end_page + 1
            ):
                sections[-1].fragments.append(fragment)
                sections[-1].end_page = page_number
                sections[-1].classification_confidence = max(
                    sections[-1].classification_confidence,
                    confidence,
                )
                continue

            if document_type == DocumentType.UNKNOWN:
                continue

            sections.append(
                PdfSection(
                    document_type=document_type,
                    classification_confidence=confidence,
                    start_page=page_number,
                    end_page=page_number,
                    section_label=section_label,
                    fragments=[fragment],
                )
            )

        return sections

    def _classify_pdf_page(
        self,
        file_name: str,
        fragment: TextFragment,
    ) -> tuple[DocumentType, float, str]:
        compact = re.sub(r"\s+", " ", fragment.text).strip().lower()
        if not compact:
            return DocumentType.UNKNOWN, 0.0, "blank page"

        if (
            "cover sheet ppa report" in compact
            or "report on production process and product approval" in compact
        ):
            return DocumentType.PSW, 0.97, "VDA/PPA cover sheet"
        if "productionprocess-related and general deliverables" in compact:
            return DocumentType.OEM_SPECIFIC, 0.86, "PPA deliverables index"
        if "product-related deliverables" in compact:
            if any(token in compact for token in ("chemical composition", "hardness:", "supplier batch no", "ts en ", "en ac ")):
                return DocumentType.MATERIAL_RESULTS, 0.95, "Material check report"
            if "actual values of organization" in compact and "requirements / specification" in compact:
                return DocumentType.DIMENSIONAL_RESULTS, 0.95, "Dimensional report"
        if compact.startswith("radiographic evaluation"):
            return DocumentType.PERFORMANCE_RESULTS, 0.95, "Radiographic evaluation"
        if "mds report" in compact or "imds id / version" in compact:
            return DocumentType.OEM_SPECIFIC, 0.9, "IMDS report"
        if "self-assessment product" in compact:
            return DocumentType.OEM_SPECIFIC, 0.9, "Self-assessment product"
        if "self-assessment process" in compact:
            return DocumentType.OEM_SPECIFIC, 0.9, "Self-assessment process"
        if "part history" in compact:
            return DocumentType.ENGINEERING_CHANGE, 0.88, "Part history"

        document_type, confidence = self._classify_document(file_name, fragment.text)
        return document_type, confidence, document_type.value.replace("_", " ").title()

    def _classify_pdf_tables(
        self,
        page_number: int,
        tables: list[TableSection],
    ) -> tuple[DocumentType, float, str]:
        page_tables = [
            table
            for table in tables
            if self._page_number_from_section_name(table.section_name) == page_number
        ]
        for table in page_tables:
            if self._is_vda_material_section(table):
                return DocumentType.MATERIAL_RESULTS, 0.96, "Material check report"
        for table in page_tables:
            if self._is_vda_dimensional_section(table):
                return DocumentType.DIMENSIONAL_RESULTS, 0.96, "Dimensional report"
        return DocumentType.UNKNOWN, 0.0, "table classification unavailable"

    def _extract_workbook(self, payload: bytes) -> tuple[list[TextFragment], list[TableSection]]:
        workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        fragments: list[TextFragment] = []
        tables: list[TableSection] = []
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            line_buffer: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                cleaned = [self._clean_cell(cell) for cell in row]
                if not any(cleaned):
                    continue
                rows.append(cleaned)
                line_buffer.append(" | ".join(cell for cell in cleaned if cell))
            if not rows:
                continue
            tables.append(TableSection(section_name=worksheet.title, rows=rows))
            fragments.append(TextFragment(text="\n".join(line_buffer), section_name=worksheet.title))
        return fragments, tables

    def _classify_document(
        self,
        file_name: str,
        text: str,
    ) -> tuple[DocumentType, float]:
        haystack = self._normalized_haystack(file_name, text)
        best_type = DocumentType.UNKNOWN
        best_signal_score = 0.0
        best_confidence = 0.28
        for document_type, keywords in DOCUMENT_KEYWORDS.items():
            score = 0.0
            hits = 0
            for token, weight in keywords:
                if self._normalized_token(token) in haystack:
                    score += weight
                    hits += 1
            if score > best_signal_score:
                best_type = document_type
                best_signal_score = score
                best_confidence = min(0.98, 0.45 + min(score, 1.6) / 2 + hits * 0.03)
        return best_type, round(best_confidence, 2)

    def _extract_metadata(
        self,
        file_name: str,
        fragments: list[TextFragment],
    ) -> dict[str, ExtractedValue]:
        metadata: dict[str, ExtractedValue] = {}
        for field_name, patterns in FIELD_PATTERNS.items():
            extracted = self._search_patterns(file_name, fragments, field_name, patterns)
            if extracted is not None:
                metadata[field_name] = extracted
        return metadata

    def _search_patterns(
        self,
        file_name: str,
        fragments: list[TextFragment],
        field_name: str,
        patterns: list[tuple[re.Pattern[str], float]],
    ) -> ExtractedValue | None:
        best_match: ExtractedValue | None = None
        best_score = -1.0
        for fragment in fragments:
            for pattern, confidence in patterns:
                for match in pattern.finditer(fragment.text):
                    value = self._clean_field_value(field_name, match.group(1))
                    if not value:
                        continue
                    snippet = self._match_snippet(fragment.text, match.span(1))
                    evidence = EvidenceRef(
                        file_name=file_name,
                        page_number=fragment.page_number,
                        section_name=fragment.section_name,
                        field_name=field_name,
                        snippet=snippet,
                        confidence=confidence,
                    )
                    score = confidence + self._field_quality_score(
                        field_name,
                        value,
                        snippet,
                        fragment.page_number,
                    )
                    if score <= best_score:
                        continue
                    best_score = score
                    best_match = ExtractedValue(
                        value=value,
                        status=ExtractionStatus.VERIFIED,
                        confidence=min(0.99, score),
                        evidence=[evidence],
                    )
        return best_match

    def _merge_filename_inference(
        self,
        file_name: str,
        metadata: dict[str, ExtractedValue],
    ) -> dict[str, ExtractedValue]:
        stem = Path(file_name).stem
        if "revision" not in metadata:
            revision_match = re.search(r"(?:^|[_\-])rev(?:ision)?[_\-]?([A-Z0-9]{1,6})(?:$|[_\-])", stem, re.IGNORECASE)
            if revision_match:
                metadata["revision"] = self._low_confidence_filename_value(
                    file_name, "revision", revision_match.group(1).upper(), 0.58
                )

        if "part_number" not in metadata:
            tokens = [token for token in re.split(r"[_\s]+", stem) if token]
            for token in tokens:
                if re.fullmatch(r"[A-Z0-9][A-Z0-9\-]{3,40}", token, re.IGNORECASE):
                    metadata["part_number"] = self._low_confidence_filename_value(
                        file_name, "part_number", token, 0.5
                    )
                    break

        return metadata

    def _merge_shared_metadata(
        self,
        metadata: dict[str, ExtractedValue],
        shared_metadata: dict[str, ExtractedValue],
    ) -> dict[str, ExtractedValue]:
        merged = dict(metadata)
        for key, value in shared_metadata.items():
            if key not in merged and value.is_present:
                merged[key] = value
        return merged

    def _low_confidence_filename_value(
        self,
        file_name: str,
        field_name: str,
        value: str,
        confidence: float,
    ) -> ExtractedValue:
        return ExtractedValue(
            value=value,
            status=ExtractionStatus.INFERRED_LOW_CONFIDENCE,
            confidence=confidence,
            evidence=[
                EvidenceRef(
                    file_name=file_name,
                    field_name=field_name,
                    snippet=f"Inferred from file name: {file_name}",
                    confidence=confidence,
                )
            ],
        )

    def _extract_structured_rows(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> dict[str, list]:
        return {
            "drawing_characteristics": self._parse_drawing_characteristics(file_name, document_type, sections),
            "inspection_results": self._parse_inspection_results(file_name, document_type, sections),
            "process_flow_steps": self._parse_process_flow_steps(file_name, document_type, sections),
            "pfmea_entries": self._parse_pfmea_entries(file_name, document_type, sections),
            "control_plan_entries": self._parse_control_plan_entries(file_name, document_type, sections),
            "certificates": self._parse_certificates(file_name, document_type, sections),
        }

    def _is_vda_dimensional_section(self, section: TableSection) -> bool:
        if self._is_vda_material_section(section):
            return False
        flattened = " ".join(
            self._clean_cell(cell).lower()
            for row in section.rows[:30]
            for cell in row
            if self._clean_cell(cell)
        )
        return (
            "requirements / specification" in flattened
            and "actual values of organization" in flattened
            and "no." in flattened
        )

    def _is_vda_material_section(self, section: TableSection) -> bool:
        flattened = " ".join(
            self._clean_cell(cell).lower()
            for row in section.rows[:35]
            for cell in row
            if self._clean_cell(cell)
        )
        if "supplier batch no" not in flattened:
            return False
        return any(
            token in flattened
            for token in ("chemical composition", "hardness:", "ts en ", "en ac ", "bronze")
        )

    def _vda_characteristic_rows(
        self,
        section: TableSection,
    ) -> tuple[list[tuple[list[str], str, list[tuple[str, str]]]], int | None]:
        if not self._is_vda_dimensional_section(section):
            return [], None

        header_row = None
        for index, row in enumerate(section.rows[:30]):
            row_text = " ".join(self._clean_cell(cell).lower() for cell in row if self._clean_cell(cell))
            if "requirements / specification" in row_text and "actual values of organization" in row_text:
                header_row = index
                break
        if header_row is None:
            return [], None

        cavity_row = section.rows[header_row + 1] if header_row + 1 < len(section.rows) else []
        measured_columns = [
            (idx, self._clean_cell(value))
            for idx, value in enumerate(cavity_row)
            if idx > 3 and self._clean_cell(value)
        ]
        if not measured_columns:
            measured_columns = [(6, "sample_1"), (10, "sample_2")]

        parsed_rows: list[tuple[list[str], str, list[tuple[str, str]]]] = []
        for row in section.rows[header_row + 1 :]:
            characteristic_id = self._clean_cell(row[0]) if row else ""
            requirement = self._clean_cell(row[3]) if len(row) > 3 else ""
            if not re.fullmatch(r"\d{2,4}", characteristic_id):
                continue
            if not requirement:
                continue
            measured_pairs: list[tuple[str, str]] = []
            for column_index, sample_label in measured_columns:
                if column_index >= len(row):
                    continue
                raw_value = self._clean_cell(row[column_index])
                if not raw_value:
                    continue
                for token_index, token in enumerate(self._split_measurement_values(raw_value), start=1):
                    label = sample_label or f"sample_{column_index}"
                    if len(self._split_measurement_values(raw_value)) > 1:
                        label = f"{label}_{token_index}"
                    measured_pairs.append((label, token))
            if measured_pairs:
                parsed_rows.append((row, requirement, measured_pairs))

        return parsed_rows, header_row

    def _split_measurement_values(self, raw_value: str) -> list[str]:
        normalized = self._clean_cell(raw_value).replace("\n", " ").strip()
        if not normalized:
            return []
        if "x" in normalized.lower():
            return [normalized]
        tokens = [token for token in re.split(r"\s+", normalized) if token]
        if len(tokens) > 1 and all(
            self._parse_numeric_value(token) is not None or token.upper() in {"OK", "GAUGE", "X"}
            for token in tokens
        ):
            return tokens
        return [normalized]

    def _derive_spec_fields(self, requirement: str) -> tuple[str | None, str | None, str | None]:
        normalized = self._normalize_spec_text(requirement).replace(",", ".")
        unit = self._guess_unit_from_spec(requirement)

        if any(keyword in normalized.lower() for keyword in ("radial run-out", "flatness", "position", "rz", "pt")):
            limit = self._extract_first_numeric(normalized)
            return None, f"<= {limit}" if limit is not None else None, unit

        nominal = self._extract_first_numeric(normalized)
        if nominal is None:
            return None, None, unit
        if "°" in normalized and any(token in normalized for token in ('"', "'", "′", "″")):
            return str(nominal), None, unit

        two_sided = re.search(
            r"\(([+-]?\d+(?:\.\d+)?)\s*/\s*([+-]?\d+(?:\.\d+)?)\)",
            normalized,
        )
        if two_sided:
            first = float(two_sided.group(1))
            second = float(two_sided.group(2))
            lower = min(first, second)
            upper = max(first, second)
            return str(nominal), f"{lower:+g}/{upper:+g}", unit

        plus_minus = re.search(r"±\s*(\d+(?:\.\d+)?)", normalized)
        if plus_minus:
            tol = float(plus_minus.group(1))
            return str(nominal), f"±{tol:g}", unit

        asym = re.search(
            r"\+(\d+(?:\.\d+)?)\s*/\s*-(\d+(?:\.\d+)?)",
            normalized,
        )
        if asym:
            return str(nominal), f"-{float(asym.group(2)):g}/+{float(asym.group(1)):g}", unit

        one_sided_plus = re.search(r"\+(\d+(?:\.\d+)?)", normalized)
        one_sided_minus = re.search(r"-(\d+(?:\.\d+)?)", normalized)
        if one_sided_plus and not one_sided_minus:
            return str(nominal), f"+0/+{float(one_sided_plus.group(1)):g}", unit
        if one_sided_minus and not one_sided_plus:
            return str(nominal), f"-{float(one_sided_minus.group(1)):g}/+0", unit

        return str(nominal), None, unit

    def _guess_unit_from_spec(self, requirement: str) -> str | None:
        normalized = self._normalize_spec_text(requirement)
        lowered = normalized.lower()
        if "°" in normalized:
            return "deg"
        if lowered.startswith("rz") or lowered.startswith("pt"):
            return "um"
        if any(token in lowered for token in ("radial run-out", "flatness", "position", "ø", "r")) or self._extract_first_numeric(normalized) is not None:
            return "mm"
        return None

    def _extract_first_numeric(self, text: str) -> float | None:
        match = re.search(r"[+-]?\d+(?:\.\d+)?", text)
        if not match:
            return None
        return float(match.group(0))

    def _parse_numeric_value(self, text: str) -> float | None:
        normalized = text.replace(",", ".")
        if "x" in normalized.lower():
            return None
        match = re.search(r"[+-]?\d+(?:\.\d+)?", normalized)
        if not match:
            return None
        return float(match.group(0))

    def _evaluate_vda_measurement(
        self,
        requirement: str,
        measured_value: str,
    ) -> MeasurementStatus:
        explicit = self._measurement_status(measured_value)
        if explicit != MeasurementStatus.UNCLEAR:
            return explicit
        if measured_value.strip().upper() in {"GAUGE", "X"}:
            return MeasurementStatus.UNCLEAR

        measured = self._parse_numeric_value(measured_value)
        if measured is None:
            return MeasurementStatus.UNCLEAR

        normalized = self._normalize_spec_text(requirement).replace(",", ".")
        lowered = normalized.lower()
        if "°" in normalized and any(token in normalized for token in ('"', "'", "′", "″")):
            return MeasurementStatus.UNCLEAR
        if any(keyword in lowered for keyword in ("radial run-out", "flatness", "position", "rz", "pt")):
            limit = self._extract_first_numeric(normalized)
            if limit is None:
                return MeasurementStatus.UNCLEAR
            return MeasurementStatus.PASS if measured <= limit + 1e-6 else MeasurementStatus.FAIL

        if "acc. to" in lowered or "pk " in lowered:
            return MeasurementStatus.UNCLEAR

        nominal = self._extract_first_numeric(normalized)
        if nominal is None:
            return MeasurementStatus.UNCLEAR

        lower_dev = upper_dev = None
        two_sided = re.search(r"\(([+-]?\d+(?:\.\d+)?)\s*/\s*([+-]?\d+(?:\.\d+)?)\)", normalized)
        if two_sided:
            first = float(two_sided.group(1))
            second = float(two_sided.group(2))
            lower_dev = min(first, second)
            upper_dev = max(first, second)
        else:
            plus_minus = re.search(r"±\s*(\d+(?:\.\d+)?)", normalized)
            asym = re.search(r"\+(\d+(?:\.\d+)?)\s*/\s*-(\d+(?:\.\d+)?)", normalized)
            if plus_minus:
                tol = float(plus_minus.group(1))
                lower_dev = -tol
                upper_dev = tol
            elif asym:
                upper_dev = float(asym.group(1))
                lower_dev = -float(asym.group(2))
            else:
                one_sided_plus = re.search(r"\+(\d+(?:\.\d+)?)", normalized)
                one_sided_minus = re.search(r"-(\d+(?:\.\d+)?)", normalized)
                if one_sided_plus and not one_sided_minus:
                    lower_dev = 0.0
                    upper_dev = float(one_sided_plus.group(1))
                elif one_sided_minus and not one_sided_plus:
                    lower_dev = -float(one_sided_minus.group(1))
                    upper_dev = 0.0

        if lower_dev is None or upper_dev is None:
            return MeasurementStatus.UNCLEAR

        lower_limit = nominal + lower_dev
        upper_limit = nominal + upper_dev
        return MeasurementStatus.PASS if lower_limit - 1e-6 <= measured <= upper_limit + 1e-6 else MeasurementStatus.FAIL

    def _parse_drawing_characteristics(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> list[DrawingCharacteristic]:
        if document_type not in {
            DocumentType.DESIGN_RECORD,
            DocumentType.BALLOONED_DRAWING,
            DocumentType.FAIR,
            DocumentType.DIMENSIONAL_RESULTS,
        }:
            return []
        characteristics: list[DrawingCharacteristic] = []
        if document_type == DocumentType.DIMENSIONAL_RESULTS:
            for section in sections:
                parsed_rows, _ = self._vda_characteristic_rows(section)
                for row, requirement, _ in parsed_rows:
                    characteristic_id = self._clean_cell(row[0])
                    nominal, tolerance, unit = self._derive_spec_fields(requirement)
                    characteristic_type = CharacteristicType.SPECIAL if any(
                        token in requirement.lower() for token in SPECIAL_TOKENS
                    ) else CharacteristicType.STANDARD
                    characteristics.append(
                        DrawingCharacteristic(
                            characteristic_id=characteristic_id,
                            balloon_number=characteristic_id,
                            description=requirement,
                            nominal=nominal,
                            tolerance=tolerance,
                            unit=unit,
                            characteristic_type=characteristic_type,
                            source_document=file_name,
                            evidence=[
                                self._row_evidence(file_name, section.section_name, "characteristic", row, 0.9)
                            ],
                        )
                    )
        for section in sections:
            header_row, mapping = self._find_header(section.rows, {"description", "nominal", "tolerance"})
            if header_row is None or not (mapping.get("characteristic_id") or mapping.get("balloon_number")):
                continue
            for row in section.rows[header_row + 1 :]:
                if not any(row):
                    continue
                characteristic_id = self._value_from_row(row, mapping, "characteristic_id") or self._value_from_row(
                    row, mapping, "balloon_number"
                )
                balloon_number = self._value_from_row(row, mapping, "balloon_number") or characteristic_id
                description = self._value_from_row(row, mapping, "description")
                if not characteristic_id or not description:
                    continue
                characteristic_type = CharacteristicType.STANDARD
                if any(token in description.lower() for token in SPECIAL_TOKENS):
                    characteristic_type = CharacteristicType.SPECIAL
                characteristics.append(
                    DrawingCharacteristic(
                        characteristic_id=characteristic_id,
                        balloon_number=balloon_number,
                        description=description,
                        nominal=self._value_from_row(row, mapping, "nominal"),
                        tolerance=self._value_from_row(row, mapping, "tolerance"),
                        unit=self._value_from_row(row, mapping, "unit"),
                        characteristic_type=characteristic_type,
                        source_document=file_name,
                        evidence=[self._row_evidence(file_name, section.section_name, "characteristic", row, 0.86)],
                    )
                )
        return characteristics

    def _parse_inspection_results(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> list[InspectionResult]:
        if document_type not in {DocumentType.DIMENSIONAL_RESULTS, DocumentType.FAIR}:
            return []
        results: list[InspectionResult] = []
        if document_type == DocumentType.DIMENSIONAL_RESULTS:
            for section in sections:
                parsed_rows, _ = self._vda_characteristic_rows(section)
                for row, requirement, measured_pairs in parsed_rows:
                    characteristic_id = self._clean_cell(row[0])
                    unit = self._guess_unit_from_spec(requirement)
                    for sample_label, measured_value in measured_pairs:
                        results.append(
                            InspectionResult(
                                characteristic_id=characteristic_id,
                                balloon_number=characteristic_id,
                                measured_value=measured_value,
                                unit=unit,
                                result=self._evaluate_vda_measurement(requirement, measured_value),
                                source_document=file_name,
                                evidence=[
                                    self._row_evidence(
                                        file_name,
                                        f"{section.section_name}:{sample_label}",
                                        "measurement",
                                        row,
                                        0.91,
                                    )
                                ],
                            )
                        )
        for section in sections:
            header_row, mapping = self._find_header(section.rows, {"measured_value"})
            if header_row is None or not (mapping.get("characteristic_id") or mapping.get("balloon_number")):
                continue
            for row in section.rows[header_row + 1 :]:
                if not any(row):
                    continue
                characteristic_id = self._value_from_row(row, mapping, "characteristic_id") or self._value_from_row(
                    row, mapping, "balloon_number"
                )
                balloon_number = self._value_from_row(row, mapping, "balloon_number")
                measured_value = self._value_from_row(row, mapping, "measured_value")
                if not characteristic_id and not balloon_number:
                    continue
                if not measured_value:
                    continue
                results.append(
                    InspectionResult(
                        characteristic_id=characteristic_id,
                        balloon_number=balloon_number,
                        measured_value=measured_value,
                        unit=self._value_from_row(row, mapping, "unit"),
                        result=self._measurement_status(self._value_from_row(row, mapping, "result")),
                        source_document=file_name,
                        evidence=[self._row_evidence(file_name, section.section_name, "measurement", row, 0.88)],
                    )
                )
        return results

    def _parse_process_flow_steps(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> list[ProcessFlowStep]:
        if document_type != DocumentType.PROCESS_FLOW:
            return []
        steps: list[ProcessFlowStep] = []
        for section in sections:
            header_row, mapping = self._find_header(section.rows, {"step_id"})
            if header_row is None:
                continue
            for sequence, row in enumerate(section.rows[header_row + 1 :], start=1):
                step_id = self._value_from_row(row, mapping, "step_id")
                if not step_id:
                    continue
                steps.append(
                    ProcessFlowStep(
                        step_id=step_id,
                        step_name=self._value_from_row(row, mapping, "step_name") or step_id,
                        sequence=sequence,
                        evidence=[self._row_evidence(file_name, section.section_name, "process_step", row, 0.84)],
                    )
                )
        return steps

    def _parse_pfmea_entries(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> list[PfmeaEntry]:
        if document_type != DocumentType.PFMEA:
            return []
        entries: list[PfmeaEntry] = []
        for section in sections:
            header_row, mapping = self._find_header(section.rows, {"step_id", "failure_mode"})
            if header_row is None:
                continue
            for row in section.rows[header_row + 1 :]:
                step_id = self._value_from_row(row, mapping, "step_id")
                failure_mode = self._value_from_row(row, mapping, "failure_mode")
                if not step_id or not failure_mode:
                    continue
                entries.append(
                    PfmeaEntry(
                        step_id=step_id,
                        failure_mode=failure_mode,
                        severity_rating=self._to_int(self._value_from_row(row, mapping, "severity")),
                        risk_priority=self._to_int(self._value_from_row(row, mapping, "risk_priority")),
                        evidence=[self._row_evidence(file_name, section.section_name, "pfmea", row, 0.82)],
                    )
                )
        return entries

    def _parse_control_plan_entries(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> list[ControlPlanEntry]:
        if document_type != DocumentType.CONTROL_PLAN:
            return []
        entries: list[ControlPlanEntry] = []
        for section in sections:
            header_row, mapping = self._find_header(section.rows, {"step_id"})
            if header_row is None or "control_method" not in mapping:
                continue
            for row in section.rows[header_row + 1 :]:
                step_id = self._value_from_row(row, mapping, "step_id")
                if not step_id:
                    continue
                characteristic_value = (
                    self._value_from_row(row, mapping, "characteristic_id")
                    or self._value_from_row(row, mapping, "balloon_number")
                    or ""
                )
                characteristic_ids = [item.strip() for item in re.split(r"[,;/]", characteristic_value) if item.strip()]
                entries.append(
                    ControlPlanEntry(
                        step_id=step_id,
                        characteristic_ids=characteristic_ids,
                        control_method=self._value_from_row(row, mapping, "control_method"),
                        reaction_plan=self._value_from_row(row, mapping, "reaction_plan"),
                        evidence=[self._row_evidence(file_name, section.section_name, "control_plan", row, 0.82)],
                    )
                )
        return entries

    def _parse_certificates(
        self,
        file_name: str,
        document_type: DocumentType,
        sections: list[TableSection],
    ) -> list[CertificateRecord]:
        if document_type not in {
            DocumentType.MATERIAL_CERTIFICATE,
            DocumentType.SPECIAL_PROCESS_CERTIFICATE,
            DocumentType.MATERIAL_RESULTS,
        }:
            return []
        records: list[CertificateRecord] = []
        seen: set[tuple[str, str | None]] = set()
        for section in sections:
            if self._is_vda_material_section(section):
                for record in self._parse_vda_material_certificates(file_name, section):
                    key = (record.certificate_type, record.identifier)
                    if key in seen:
                        continue
                    seen.add(key)
                    records.append(record)
            header_row, mapping = self._find_header(section.rows, {"identifier"})
            if header_row is None:
                continue
            for row in section.rows[header_row + 1 :]:
                identifier = self._value_from_row(row, mapping, "identifier")
                if not identifier:
                    continue
                record = CertificateRecord(
                    certificate_type=self._value_from_row(row, mapping, "certificate_type") or document_type.value,
                    identifier=identifier,
                    source_document=file_name,
                    evidence=[self._row_evidence(file_name, section.section_name, "certificate", row, 0.8)],
                )
                key = (record.certificate_type, record.identifier)
                if key in seen:
                    continue
                seen.add(key)
                records.append(record)
        return records

    def _parse_vda_material_certificates(
        self,
        file_name: str,
        section: TableSection,
    ) -> list[CertificateRecord]:
        if not self._is_vda_material_section(section):
            return []

        records: list[CertificateRecord] = []
        header_row = None
        for index, row in enumerate(section.rows[:30]):
            row_text = " ".join(self._clean_cell(cell).lower() for cell in row if self._clean_cell(cell))
            if "requirements / specification" in row_text and "actual values of organization" in row_text:
                header_row = index
                break
        if header_row is None:
            return []

        for row in section.rows[header_row + 1 :]:
            requirement = self._clean_cell(row[3]) if len(row) > 3 else ""
            actual_value = self._clean_cell(row[6]) if len(row) > 6 else ""
            if not requirement:
                continue

            lowered = requirement.lower()
            if lowered == "chemical composition (%)":
                break

            if lowered.startswith("supplier batch no"):
                if actual_value and actual_value != "-":
                    records.append(
                        CertificateRecord(
                            certificate_type="supplier_batch",
                            identifier=actual_value,
                            related_requirement=requirement,
                            source_document=file_name,
                            evidence=[self._row_evidence(file_name, section.section_name, "supplier_batch", row, 0.88)],
                        )
                    )
                continue

            if requirement.upper().startswith(("TS EN", "EN AC", "DIN", "ISO", "ASTM")) or "bronze" in lowered:
                records.append(
                    CertificateRecord(
                        certificate_type="material_specification",
                        identifier=requirement,
                        related_requirement=actual_value or None,
                        source_document=file_name,
                        evidence=[self._row_evidence(file_name, section.section_name, "material_specification", row, 0.86)],
                    )
                )
        return records

    def _find_header(
        self,
        rows: list[list[str]],
        required_fields: set[str],
    ) -> tuple[int | None, dict[str, int]]:
        for index, row in enumerate(rows[:25]):
            mapping = self._header_mapping(row)
            if required_fields.issubset(mapping.keys()):
                return index, mapping
        return None, {}

    def _header_mapping(self, row: list[str]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for index, cell in enumerate(row):
            normalized = self._normalize_header(cell)
            if not normalized:
                continue
            for canonical, aliases in HEADER_SYNONYMS.items():
                if canonical in mapping:
                    continue
                if normalized == canonical or normalized in aliases:
                    mapping[canonical] = index
                    break
        return mapping

    def _value_from_row(
        self,
        row: list[str],
        mapping: dict[str, int],
        field_name: str,
    ) -> str | None:
        index = mapping.get(field_name)
        if index is None or index >= len(row):
            return None
        value = self._clean_cell(row[index])
        return value or None

    def _hydrate_context(
        self,
        context: SubmissionContext,
        documents: Iterable[DocumentRecord],
    ) -> SubmissionContext:
        updates: dict[str, object] = {}
        if context.ppap_level is None:
            for document in documents:
                field = document.metadata.get("ppap_level")
                if field and field.is_present and field.text.isdigit():
                    updates["ppap_level"] = int(field.text)
                    break

        for context_key, metadata_key in (
            ("part_number", "part_number"),
            ("drawing_number", "drawing_number"),
            ("revision", "revision"),
            ("customer_oem", "customer_name"),
            ("supplier_name", "supplier_name"),
            ("manufacturing_process", "process_name"),
            ("material", "material"),
        ):
            if getattr(context, context_key):
                continue
            for document in documents:
                field = document.metadata.get(metadata_key)
                if field and field.is_present:
                    updates[context_key] = field.text
                    break

        return context.model_copy(update=updates)

    def _submission_id(self) -> str:
        timestamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S")
        return f"upload-{timestamp}-{uuid.uuid4().hex[:8]}"

    def _measurement_status(self, value: str | None) -> MeasurementStatus:
        if not value:
            return MeasurementStatus.UNCLEAR
        normalized = value.strip().lower()
        if normalized in PASS_TOKENS:
            return MeasurementStatus.PASS
        if normalized in FAIL_TOKENS:
            return MeasurementStatus.FAIL
        return MeasurementStatus.UNCLEAR

    def _normalize_capture(self, value: str) -> str:
        value = re.sub(r"\s+", " ", value.replace("\u00a0", " ")).strip(" :;-")
        return value

    def _clean_field_value(self, field_name: str, value: str) -> str | None:
        cleaned = self._normalize_capture(value)
        if not cleaned:
            return None

        if field_name in {"part_number", "drawing_number"}:
            cleaned = re.split(
                r"\b(?:hardware version|part number|name|drawing number|diagnosis status|software version|version / date|identification(?:/duns)?)\b",
                cleaned,
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0].strip(" -")
            if cleaned.upper() in {"NO", "NO.", "YES", "OK", "NOK", "-"}:
                return None
            if len(re.sub(r"[^A-Z0-9]", "", cleaned.upper())) < 4:
                return None

        if field_name == "revision":
            cleaned = cleaned.split("/")[0].strip()
            if not cleaned or cleaned.lower().startswith("fr."):
                return None
            if any(token in cleaned.lower() for token in MONTH_TOKENS):
                return None
            if len(cleaned) > 12:
                return None

        if field_name == "customer_name":
            if "customer" in cleaned.lower() and not cleaned.lower().startswith("customer"):
                cleaned = cleaned.lower().rsplit("customer", 1)[-1].strip(" -").title()
            cleaned = re.split(
                r"\b(?:report number|delivery note number|report version|delivery quantity|delivery location|batch number|order number|production location|sample weight|unloading point|remark|department|telephone|signature|date)\b",
                cleaned,
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0].strip(" -")
            if any(
                token in cleaned.lower()
                for token in (
                    "customer-ready",
                    "not customer-ready",
                    "ready for series production",
                    "decision",
                    "new ppa procedure required",
                    "update of ppa documentation required",
                    "new part",
                )
            ):
                return None

        if field_name == "supplier_name":
            cleaned = re.split(
                r"\b(?:organization reason for report creation|information about the organization|information about samples|information about the customer)\b",
                cleaned,
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0].strip(" -")
            if any(
                token in cleaned.lower()
                for token in ("report number", "delivery note", "part number", "code:", "change number")
            ):
                return None
            if cleaned.lower() in {"version", "customer", "organization", "name"}:
                return None

        if field_name == "process_name" and cleaned.lower() in {
            "related and general",
            "product-related",
            "process-related and general",
            "productionprocess-related and general",
        }:
            return None

        if field_name == "material":
            if cleaned.upper() in {"NO", "NO.", "YES", "OK", "NOK"}:
                return None
            if any(
                token in cleaned.lower()
                for token in (
                    "series material",
                    "materials which are subject",
                    "specification not met",
                    "customer acceptance granted",
                )
            ):
                return None

        return cleaned or None

    def _field_quality_score(
        self,
        field_name: str,
        value: str,
        snippet: str,
        page_number: int | None,
    ) -> float:
        score = 0.0
        lowered_value = value.lower()
        lowered_snippet = snippet.lower()

        if page_number and page_number <= 3:
            score += 0.03
        if field_name in {"part_number", "drawing_number"} and any(char.isdigit() for char in value):
            score += 0.03
        if field_name == "supplier_name" and any(token in lowered_snippet for token in ("organization", "supplier")):
            score += 0.05
        if field_name == "customer_name" and "customer" in lowered_snippet:
            score += 0.04
        if field_name == "revision" and re.fullmatch(r"[A-Z0-9]{1,8}", value):
            score += 0.04
        if field_name == "process_name":
            score -= 0.05
        if any(token in lowered_value for token in ("report version", "delivery quantity", "telephone")):
            score -= 0.08
        return score

    def _normalized_haystack(self, file_name: str, text: str) -> str:
        normalized = re.sub(r"[^a-z0-9]+", " ", f"{file_name} {text}".lower())
        return f" {normalized.strip()} "

    def _normalized_token(self, token: str) -> str:
        normalized = re.sub(r"[^a-z0-9]+", " ", token.lower()).strip()
        return f" {normalized} "

    def _match_snippet(self, text: str, span: tuple[int, int]) -> str:
        start, end = span
        left = max(0, start - 40)
        right = min(len(text), end + 60)
        return self._normalize_capture(text[left:right])

    def _clean_cell(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    def _normalize_spec_text(self, value: str) -> str:
        return (
            value.replace("Â±", "±")
            .replace("Ã¸", "ø")
            .replace("Ø", "ø")
            .replace("Â°", "°")
        )

    def _normalize_header(self, value: str) -> str:
        value = self._clean_cell(value).lower()
        return re.sub(r"[^a-z0-9]+", " ", value).strip()

    def _to_int(self, value: str | None) -> int | None:
        if value is None:
            return None
        match = re.search(r"-?\d+", value)
        if not match:
            return None
        return int(match.group(0))

    def _row_evidence(
        self,
        file_name: str,
        section_name: str,
        field_name: str,
        row: list[str],
        confidence: float,
    ) -> EvidenceRef:
        return EvidenceRef(
            file_name=file_name,
            section_name=section_name,
            field_name=field_name,
            snippet=" | ".join(cell for cell in row if cell)[:220],
            confidence=confidence,
        )

    def _section_file_name(self, file_name: str, start_page: int, end_page: int) -> str:
        return f"{file_name} [{self._page_span_text(start_page, end_page)}]"

    def _page_span_text(self, start_page: int, end_page: int) -> str:
        if start_page == end_page:
            return f"p. {start_page}"
        return f"pp. {start_page}-{end_page}"

    def _tables_for_page_span(
        self,
        tables: list[TableSection],
        start_page: int,
        end_page: int,
    ) -> list[TableSection]:
        selected: list[TableSection] = []
        for table in tables:
            page_number = self._page_number_from_section_name(table.section_name)
            if page_number is None:
                continue
            if start_page <= page_number <= end_page:
                selected.append(table)
        return selected

    def _page_number_from_section_name(self, section_name: str) -> int | None:
        match = re.search(r"pdf_page_(\d+)", section_name)
        if not match:
            return None
        return int(match.group(1))
